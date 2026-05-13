from django.http import HttpResponse
from django.template.loader import render_to_string
from io import BytesIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from xhtml2pdf import pisa
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom


# ============================================================================
# PDF GENERATION - usando xhtml2pdf (OBBLIGATORIO)
# ============================================================================


def genera_pdf_da_template(template_name, context, filename="document.pdf"):
    """
    Genera un PDF da un template HTML usando xhtml2pdf.

    Args:
        template_name: path del template HTML (es. 'pdf/fattura.html')
        context: dizionario con i dati da passare al template
        filename: nome del file PDF da scaricare

    Returns:
        HttpResponse con il PDF generato

    Esempio:
        context = {'ordine': ordine, 'cliente': cliente}
        return genera_pdf_da_template('pdf/ordine.html', context, 'ordine_123.pdf')
    """
    # Renderizza il template HTML
    html_string = render_to_string(template_name, context)

    # Crea il PDF in memoria
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result)

    if pdf.err:
        return HttpResponse("Errore durante la generazione del PDF", status=500)

    # Restituisce il PDF come response
    response = HttpResponse(result.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def genera_pdf_da_html(html_content, filename="document.pdf"):
    """
    Genera un PDF direttamente da una stringa HTML.

    Args:
        html_content: stringa HTML da convertire
        filename: nome del file PDF

    Returns:
        HttpResponse con il PDF generato
    """
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)

    if pdf.err:
        return HttpResponse("Errore durante la generazione del PDF", status=500)

    response = HttpResponse(result.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ============================================================================
# EXCEL GENERATION - usando openpyxl
# ============================================================================


def genera_excel_da_queryset(
    queryset,
    columns,
    filename="export.xlsx",
    sheet_name="Dati",
    include_header=True,
    header_style=True,
):
    """
    Genera un file Excel da un queryset Django.

    Args:
        queryset: queryset Django da esportare
        columns: lista di tuple (nome_campo, intestazione_colonna)
                 es. [('id', 'ID'), ('nome', 'Nome Cliente')]
        filename: nome del file Excel
        sheet_name: nome del foglio
        include_header: se True, include la riga di intestazione
        header_style: se True, applica stile all'intestazione

    Returns:
        HttpResponse con il file Excel

    Esempio:
        columns = [
            ('numero', 'Numero Ordine'),
            ('cliente__ragione_sociale', 'Cliente'),
            ('data', 'Data'),
            ('totale', 'Totale €')
        ]
        return genera_excel_da_queryset(
            Ordine.objects.all(),
            columns,
            'ordini_export.xlsx'
        )
    """
    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Stile header
    header_fill = PatternFill(
        start_color="5585b5", end_color="5585b5", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Scrivi header
    if include_header:
        for col_idx, (field, header_text) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            if header_style:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

    # Scrivi dati
    start_row = 2 if include_header else 1
    for row_idx, obj in enumerate(queryset, start=start_row):
        for col_idx, (field, _) in enumerate(columns, start=1):
            # Supporta campi relazionali con __
            value = obj
            for field_part in field.split("__"):
                value = getattr(value, field_part, "")
                if callable(value):
                    value = value()

            ws.cell(row=row_idx, column=col_idx, value=str(value))

    # Auto-size colonne
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Salva in memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Restituisce come response
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def genera_excel_personalizzato(data_dict, filename="export.xlsx"):
    """
    Genera un Excel con più fogli personalizzati.

    Args:
        data_dict: dizionario con struttura:
            {
                'Foglio1': {
                    'headers': ['Col1', 'Col2'],
                    'data': [[val1, val2], [val3, val4]]
                },
                'Foglio2': {...}
            }
        filename: nome del file

    Returns:
        HttpResponse con il file Excel
    """
    wb = Workbook()
    wb.remove(wb.active)  # Rimuovi foglio default

    header_fill = PatternFill(
        start_color="5585b5", end_color="5585b5", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for sheet_name, sheet_data in data_dict.items():
        ws = wb.create_sheet(title=sheet_name)

        # Headers
        if "headers" in sheet_data:
            for col_idx, header in enumerate(sheet_data["headers"], start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font

        # Data
        if "data" in sheet_data:
            for row_idx, row_data in enumerate(sheet_data["data"], start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ============================================================================
# CSV GENERATION
# ============================================================================


def genera_csv_da_queryset(queryset, columns, filename="export.csv"):
    """
    Genera un file CSV da un queryset Django.

    Args:
        queryset: queryset Django
        columns: lista di tuple (campo, intestazione)
        filename: nome del file CSV

    Returns:
        HttpResponse con il file CSV

    Esempio:
        columns = [('numero', 'Numero'), ('cliente__nome', 'Cliente')]
        return genera_csv_da_queryset(ordini, columns, 'ordini.csv')
    """
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # BOM per Excel compatibilità UTF-8
    response.write("\ufeff")

    writer = csv.writer(response, delimiter=";")

    # Header
    writer.writerow([header for _, header in columns])

    # Dati
    for obj in queryset:
        row = []
        for field, _ in columns:
            value = obj
            for field_part in field.split("__"):
                value = getattr(value, field_part, "")
                if callable(value):
                    value = value()
            row.append(str(value))
        writer.writerow(row)

    return response


# ============================================================================
# XML FATTURA ELETTRONICA - Formato SDI (Sistema di Interscambio)
# ============================================================================


def genera_xml_fattura_elettronica(fattura_data, filename="fattura.xml"):
    """
    Genera XML per Fattura Elettronica formato SDI conforme specifiche Agenzia Entrate.

    Versione: FPR12 (Fattura PA 1.2)
    Conforme a: Specifiche tecniche Fatturazione Elettronica v1.7.1

    ⚠️ IMPORTANTE:
    - Questa funzione genera XML strutturalmente corretto
    - Per produzione serve validazione XSD e firma digitale
    - Verificare sempre conformità con specifiche aggiornate
    - Controllare partite IVA e codici fiscali

    Args:
        fattura_data: dizionario con i dati della fattura
        filename: nome del file XML

    Returns:
        HttpResponse con il file XML

    Struttura fattura_data esempio:
    {
        'cedente': {
            'paese': 'IT',
            'partita_iva': '12345678901',  # Senza IT
            'codice_fiscale': '12345678901',
            'denominazione': 'Azienda SRL',
            'regime_fiscale': 'RF01',  # Ordinario
            'indirizzo': 'Via Roma 1',
            'numero_civico': '1',
            'cap': '00100',
            'comune': 'Roma',
            'provincia': 'RM',
            'nazione': 'IT'
        },
        'cessionario': {
            'paese': 'IT',
            'id_codice': '87654321098',  # P.IVA o CF
            'codice_fiscale': '87654321098',  # Opzionale
            'denominazione': 'Cliente SRL',  # Oppure 'nome' e 'cognome' per persona fisica
            'indirizzo': 'Via Milano 10',
            'numero_civico': '10',
            'cap': '20100',
            'comune': 'Milano',
            'provincia': 'MI',
            'nazione': 'IT'
        },
        'numero': '001',
        'data': '2025-01-15',
        'tipo_documento': 'TD01',  # Fattura
        'divisa': 'EUR',
        'codice_destinatario': '0000000',  # 7 caratteri, o PEC
        'causale': 'Vendita merce',  # Opzionale
        'linee': [
            {
                'numero': 1,
                'descrizione': 'Prodotto 1',
                'quantita': 10.00,
                'unita_misura': 'PZ',  # Opzionale
                'prezzo_unitario': 50.00,
                'aliquota_iva': 22.00,
                'natura': None  # Se esente IVA: 'N1', 'N2', etc.
            }
        ],
        'pagamento': {
            'condizioni': 'TP02',  # Pagamento completo
            'modalita': 'MP05',  # Bonifico
            'importo': 610.00,  # Calcolato automaticamente se non fornito
            'data_scadenza': '2025-02-15',  # Opzionale
            'iban': 'IT60X0542811101000000123456'  # Opzionale
        }
    }
    """
    # Root element
    root = Element("p:FatturaElettronica")
    root.set("xmlns:ds", "http://www.w3.org/2000/09/xmldsig#")
    root.set("xmlns:p", "http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2")
    root.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
    root.set("versione", "FPR12")
    root.set(
        "xsi:schemaLocation",
        "http://ivaservizi.agenziaentrate.gov.it/docs/xsd/fatture/v1.2 http://www.fatturapa.gov.it/export/fatturazione/sdi/fatturapa/v1.2/Schema_del_file_xml_FatturaPA_versione_1.2.xsd",
    )

    # ========== HEADER ==========
    header = SubElement(root, "FatturaElettronicaHeader")

    # --- Dati Trasmissione ---
    dati_trasm = SubElement(header, "DatiTrasmissione")
    id_trasm = SubElement(dati_trasm, "IdTrasmittente")
    SubElement(id_trasm, "IdPaese").text = fattura_data["cedente"].get("paese", "IT")
    SubElement(id_trasm, "IdCodice").text = fattura_data["cedente"]["partita_iva"]
    SubElement(dati_trasm, "ProgressivoInvio").text = fattura_data["numero"]
    SubElement(dati_trasm, "FormatoTrasmissione").text = "FPR12"
    SubElement(dati_trasm, "CodiceDestinatario").text = fattura_data.get(
        "codice_destinatario", "0000000"
    )

    # --- Cedente Prestatore ---
    cedente = SubElement(header, "CedentePrestatore")
    dati_anag_cedente = SubElement(cedente, "DatiAnagrafici")

    # IdFiscaleIVA
    id_fiscale_iva_cedente = SubElement(dati_anag_cedente, "IdFiscaleIVA")
    SubElement(id_fiscale_iva_cedente, "IdPaese").text = fattura_data["cedente"].get(
        "paese", "IT"
    )
    SubElement(id_fiscale_iva_cedente, "IdCodice").text = fattura_data["cedente"][
        "partita_iva"
    ]

    # Codice Fiscale (se diverso da P.IVA)
    if "codice_fiscale" in fattura_data["cedente"]:
        SubElement(dati_anag_cedente, "CodiceFiscale").text = fattura_data["cedente"][
            "codice_fiscale"
        ]

    # Anagrafica
    anagrafica_cedente = SubElement(dati_anag_cedente, "Anagrafica")
    if "denominazione" in fattura_data["cedente"]:
        SubElement(anagrafica_cedente, "Denominazione").text = fattura_data["cedente"][
            "denominazione"
        ]
    else:
        # Persona fisica
        SubElement(anagrafica_cedente, "Nome").text = fattura_data["cedente"].get(
            "nome", ""
        )
        SubElement(anagrafica_cedente, "Cognome").text = fattura_data["cedente"].get(
            "cognome", ""
        )

    # Regime Fiscale
    SubElement(dati_anag_cedente, "RegimeFiscale").text = fattura_data["cedente"].get(
        "regime_fiscale", "RF01"
    )

    # Sede Cedente
    sede_cedente = SubElement(cedente, "Sede")
    SubElement(sede_cedente, "Indirizzo").text = fattura_data["cedente"]["indirizzo"]
    if "numero_civico" in fattura_data["cedente"]:
        SubElement(sede_cedente, "NumeroCivico").text = str(
            fattura_data["cedente"]["numero_civico"]
        )
    SubElement(sede_cedente, "CAP").text = fattura_data["cedente"]["cap"]
    SubElement(sede_cedente, "Comune").text = fattura_data["cedente"]["comune"]
    if "provincia" in fattura_data["cedente"]:
        SubElement(sede_cedente, "Provincia").text = fattura_data["cedente"][
            "provincia"
        ]
    SubElement(sede_cedente, "Nazione").text = fattura_data["cedente"].get(
        "nazione", "IT"
    )

    # --- Cessionario Committente ---
    cessionario = SubElement(header, "CessionarioCommittente")
    dati_anag_cess = SubElement(cessionario, "DatiAnagrafici")

    # IdFiscaleIVA o CodiceFiscale
    if (
        "partita_iva" in fattura_data["cessionario"]
        and fattura_data["cessionario"]["partita_iva"]
    ):
        id_fiscale_iva_cess = SubElement(dati_anag_cess, "IdFiscaleIVA")
        SubElement(id_fiscale_iva_cess, "IdPaese").text = fattura_data[
            "cessionario"
        ].get("paese", "IT")
        SubElement(id_fiscale_iva_cess, "IdCodice").text = fattura_data["cessionario"][
            "partita_iva"
        ]

    if "codice_fiscale" in fattura_data["cessionario"]:
        SubElement(dati_anag_cess, "CodiceFiscale").text = fattura_data["cessionario"][
            "codice_fiscale"
        ]
    elif "id_codice" in fattura_data["cessionario"]:
        # Fallback per compatibilità
        SubElement(dati_anag_cess, "CodiceFiscale").text = fattura_data["cessionario"][
            "id_codice"
        ]

    # Anagrafica Cessionario
    anagrafica_cess = SubElement(dati_anag_cess, "Anagrafica")
    if "denominazione" in fattura_data["cessionario"]:
        SubElement(anagrafica_cess, "Denominazione").text = fattura_data["cessionario"][
            "denominazione"
        ]
    else:
        SubElement(anagrafica_cess, "Nome").text = fattura_data["cessionario"].get(
            "nome", ""
        )
        SubElement(anagrafica_cess, "Cognome").text = fattura_data["cessionario"].get(
            "cognome", ""
        )

    # Sede Cessionario
    sede_cess = SubElement(cessionario, "Sede")
    SubElement(sede_cess, "Indirizzo").text = fattura_data["cessionario"]["indirizzo"]
    if "numero_civico" in fattura_data["cessionario"]:
        SubElement(sede_cess, "NumeroCivico").text = str(
            fattura_data["cessionario"]["numero_civico"]
        )
    SubElement(sede_cess, "CAP").text = fattura_data["cessionario"]["cap"]
    SubElement(sede_cess, "Comune").text = fattura_data["cessionario"]["comune"]
    if "provincia" in fattura_data["cessionario"]:
        SubElement(sede_cess, "Provincia").text = fattura_data["cessionario"][
            "provincia"
        ]
    SubElement(sede_cess, "Nazione").text = fattura_data["cessionario"].get(
        "nazione", "IT"
    )

    # ========== BODY ==========
    body = SubElement(root, "FatturaElettronicaBody")

    # --- Dati Generali ---
    dati_gen = SubElement(body, "DatiGenerali")
    dati_gen_doc = SubElement(dati_gen, "DatiGeneraliDocumento")
    SubElement(dati_gen_doc, "TipoDocumento").text = fattura_data.get(
        "tipo_documento", "TD01"
    )
    SubElement(dati_gen_doc, "Divisa").text = fattura_data.get("divisa", "EUR")
    SubElement(dati_gen_doc, "Data").text = fattura_data["data"]
    SubElement(dati_gen_doc, "Numero").text = fattura_data["numero"]

    # Causale (opzionale)
    if "causale" in fattura_data and fattura_data["causale"]:
        SubElement(dati_gen_doc, "Causale").text = fattura_data["causale"][
            :200
        ]  # Max 200 caratteri

    # --- Calcola totali e raggruppa per aliquota IVA ---
    totali_per_aliquota = {}
    imponibile_totale = 0
    iva_totale = 0

    # --- Linee Dettaglio ---
    dati_beni = SubElement(body, "DatiBeniServizi")
    for linea in fattura_data.get("linee", []):
        dettaglio = SubElement(dati_beni, "DettaglioLinee")
        SubElement(dettaglio, "NumeroLinea").text = str(linea["numero"])
        SubElement(dettaglio, "Descrizione").text = linea["descrizione"][
            :1000
        ]  # Max 1000 caratteri

        quantita = float(linea["quantita"])
        prezzo_unitario = float(linea["prezzo_unitario"])

        SubElement(dettaglio, "Quantita").text = f"{quantita:.2f}"
        if "unita_misura" in linea and linea["unita_misura"]:
            SubElement(dettaglio, "UnitaMisura").text = linea["unita_misura"]
        SubElement(dettaglio, "PrezzoUnitario").text = f"{prezzo_unitario:.8f}"

        imponibile_linea = quantita * prezzo_unitario
        SubElement(dettaglio, "PrezzoTotale").text = f"{imponibile_linea:.2f}"

        # Gestione IVA o Natura (se esente)
        if linea.get("natura"):
            # Operazione esente/fuori campo IVA
            SubElement(dettaglio, "AliquotaIVA").text = "0.00"
            SubElement(dettaglio, "Natura").text = linea["natura"]
        else:
            # Operazione con IVA
            aliquota = float(linea["aliquota_iva"])
            SubElement(dettaglio, "AliquotaIVA").text = f"{aliquota:.2f}"

            # Accumula per aliquota
            if aliquota not in totali_per_aliquota:
                totali_per_aliquota[aliquota] = {
                    "imponibile": 0,
                    "imposta": 0,
                    "natura": None,
                }

            totali_per_aliquota[aliquota]["imponibile"] += imponibile_linea
            totali_per_aliquota[aliquota]["imposta"] += imponibile_linea * (
                aliquota / 100
            )

        imponibile_totale += imponibile_linea

    # --- Dati Riepilogo (uno per ogni aliquota IVA) ---
    for aliquota, valori in sorted(totali_per_aliquota.items()):
        riepilogo = SubElement(dati_beni, "DatiRiepilogo")
        SubElement(riepilogo, "AliquotaIVA").text = f"{aliquota:.2f}"
        if valori["natura"]:
            SubElement(riepilogo, "Natura").text = valori["natura"]
        SubElement(riepilogo, "ImponibileImporto").text = f"{valori['imponibile']:.2f}"
        SubElement(riepilogo, "Imposta").text = f"{valori['imposta']:.2f}"
        SubElement(riepilogo, "EsigibilitaIVA").text = "I"  # Immediata
        iva_totale += valori["imposta"]

    # --- Dati Pagamento ---
    if "pagamento" in fattura_data:
        pag_data = fattura_data["pagamento"]
        dati_pag = SubElement(body, "DatiPagamento")
        SubElement(dati_pag, "CondizioniPagamento").text = pag_data.get(
            "condizioni", "TP02"
        )

        dettaglio_pag = SubElement(dati_pag, "DettaglioPagamento")
        SubElement(dettaglio_pag, "ModalitaPagamento").text = pag_data.get(
            "modalita", "MP05"
        )

        # Importo (calcolato se non fornito)
        importo_pagamento = pag_data.get("importo", imponibile_totale + iva_totale)
        SubElement(dettaglio_pag, "ImportoPagamento").text = f"{importo_pagamento:.2f}"

        # Data scadenza (opzionale)
        if "data_scadenza" in pag_data:
            SubElement(dettaglio_pag, "DataScadenzaPagamento").text = pag_data[
                "data_scadenza"
            ]

        # IBAN (opzionale)
        if "iban" in pag_data and pag_data["iban"]:
            SubElement(dettaglio_pag, "IBAN").text = pag_data["iban"]

    # Converti in stringa XML formattata
    xml_string = minidom.parseString(tostring(root)).toprettyxml(indent="  ")

    # Restituisce come response
    response = HttpResponse(xml_string, content_type="application/xml")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ============================================================================
# UTILITY GENERICHE
# ============================================================================


def get_export_response(data, export_format, filename_base, **kwargs):
    """
    Utility generica per esportare dati in diversi formati.

    Args:
        data: queryset o lista di dati
        export_format: 'pdf', 'excel', 'csv', 'xml'
        filename_base: nome base del file senza estensione
        **kwargs: parametri aggiuntivi specifici per formato

    Returns:
        HttpResponse appropriato per il formato richiesto
    """
    if export_format == "pdf":
        template = kwargs.get("template_name")
        context = kwargs.get("context", {})
        return genera_pdf_da_template(template, context, f"{filename_base}.pdf")

    elif export_format == "excel":
        columns = kwargs.get("columns", [])
        return genera_excel_da_queryset(data, columns, f"{filename_base}.xlsx")

    elif export_format == "csv":
        columns = kwargs.get("columns", [])
        return genera_csv_da_queryset(data, columns, f"{filename_base}.csv")

    elif export_format == "xml":
        fattura_data = kwargs.get("fattura_data", {})
        return genera_xml_fattura_elettronica(fattura_data, f"{filename_base}.xml")

    else:
        return HttpResponse("Formato non supportato", status=400)


def serve_qr_code(request):
    """
    Vista per generare e servire un QR Code on-the-fly tramite query parameter.
    
    Esempio: /core/qrcode/?data=https://google.com&size=10
    """
    from .qr_code_generator import generate_qr_code

    data = request.GET.get("data")
    if not data:
        return HttpResponse("Parametro 'data' mancante", status=400)

    try:
        box_size = int(request.GET.get("size", 10))
        border = int(request.GET.get("border", 4))
    except ValueError:
        return HttpResponse("Parametri size o border non validi", status=400)

    qr_buffer = generate_qr_code(data, box_size=box_size, border=border)
    return HttpResponse(qr_buffer.getvalue(), content_type="image/png")
