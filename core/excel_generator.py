"""
CORE EXCEL GENERATOR - ModularBEF
==================================

Sistema universale per generazione file Excel.
Supporta openpyxl per Excel professionale con stili e formule.
"""

from io import BytesIO
from typing import List, Dict, Any
from datetime import datetime, date
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def generate_excel_response(
    data: List[Dict[str, Any]],
    filename: str,
    sheet_name: str = "Dati",
    headers: List[str] = None,
) -> HttpResponse:
    """
    Genera un file Excel e lo ritorna come HttpResponse.

    Args:
        data: Lista di dizionari con i dati
        filename: Nome del file (senza estensione)
        sheet_name: Nome del foglio
        headers: Lista headers personalizzati (opzionale)

    Returns:
        HttpResponse con il file Excel
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl non installato. Run: pip install openpyxl")

    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        # File vuoto
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response

    # Headers
    if headers is None:
        headers = list(data[0].keys())

    # Stile header
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(
        start_color="5585b5", end_color="5585b5", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Scrivi headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Scrivi dati
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            value = row_data.get(header, "")

            # Formatta valore
            if isinstance(value, (datetime, date)):
                cell.value = (
                    value.strftime("%d/%m/%Y %H:%M")
                    if isinstance(value, datetime)
                    else value.strftime("%d/%m/%Y")
                )
            elif isinstance(value, Decimal):
                cell.value = float(value)
            else:
                cell.value = value

            # Allineamento
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-width colonne
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Bordi
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in ws.iter_rows(
        min_row=1, max_row=len(data) + 1, min_col=1, max_col=len(headers)
    ):
        for cell in row:
            cell.border = thin_border

    # Freeze panes (blocca riga header)
    ws.freeze_panes = "A2"

    # Salva in response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'

    return response
